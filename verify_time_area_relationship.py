"""Regression checks for effective-exploration-time versus total coverage."""

from __future__ import annotations

import math
from pathlib import Path

from statistical_analysis import analyze_results, write_results_excel


def _rows(kind: str) -> list[dict]:
    rows = []
    times = (60, 90, 120, 160, 210, 270, 340, 420, 510, 600)
    for index, texpl in enumerate(times):
        jitter = 0.20 if index % 2 else -0.20
        if kind == "linear":
            atot = 8.0 + 0.09 * texpl + jitter
        elif kind == "saturating":
            atot = 10.0 + 75.0 * (1.0 - math.exp(-texpl / 180.0)) + jitter
        else:  # pragma: no cover - test authoring error
            raise ValueError(kind)
        rows.append({
            "run": index + 1,
            "start_attempt": 0,
            "diagnostic_restarts": 0,
            "Fw": 0.0,
            "Opt": bool(index % 2),
            "Iw": 1.0,
            "Dw": 1.0,
            "Vw": 0.3,
            "Cw": 3.0,
            "Ow": 8.0,
            "Wp": 24.0,
            "Ms": 1.0,
            "layout_mode": "office",
            "include_objects": True,
            "object_density": 1.0,
            "Af": atot,
            "Af_star": atot,
            "Atot": atot,
            "Vf": 2,
            "Cf": 1,
            "texpl": float(texpl),
            "TP": 1,
            "FP": 0,
            "FN": 0,
            "TN": 10.0,
            "SAR_Sensitivity": 1.0,
            "SAR_Specificity": 1.0,
            "SAR_BalancedAccuracy": 1.0,
            "SAR_MCC": 1.0,
        })
    return rows


def _summary(rows: list[dict]) -> list[dict]:
    return [{
        "Fw": 0.0,
        "Opt": False,
        "N": len(rows),
        "Af_mean": sum(row["Af"] for row in rows) / len(rows),
        "Af_std": 0.0,
        "Af_star_mean": sum(row["Af_star"] for row in rows) / len(rows),
        "Af_star_std": 0.0,
        "Atot_mean": sum(row["Atot"] for row in rows) / len(rows),
        "Atot_std": 0.0,
    }]


def main() -> None:
    linear_rows = _rows("linear")
    linear_analysis = analyze_results(linear_rows)
    linear = linear_analysis["time_area_relationship"][0]
    assert linear["scope"] == "Complessiva"
    assert linear["relationship"] == "approssimativamente lineare"
    assert linear["best_model"] == "linear"
    assert linear["linear_r2"] > 0.99
    assert linear["linear_slope_p"] < 0.001

    saturating_rows = _rows("saturating")
    saturating_analysis = analyze_results(saturating_rows)
    saturating = saturating_analysis["time_area_relationship"][0]
    assert saturating["relationship"].startswith("non lineare")
    assert saturating["best_model"] in {
        "logarithmic", "square_root", "quadratic", "saturating_exponential"
    }
    assert saturating["delta_aicc_vs_linear"] >= 2.0
    assert saturating["nonlinearity_evidence"] in {"debole", "moderata", "forte"}

    output = Path("_verify_v38_time_area.xlsx")
    write_results_excel(
        output, saturating_rows, _summary(saturating_rows), saturating_analysis)
    assert output.exists() and output.stat().st_size > 0

    from openpyxl import load_workbook
    workbook = load_workbook(output, read_only=True)
    assert "Time-area relation" in workbook.sheetnames
    assert "Time-area models" in workbook.sheetnames
    relation_headers = [
        cell.value for cell in next(workbook["Time-area relation"].iter_rows())
    ]
    assert "relationship" in relation_headers
    assert "delta_aicc_vs_linear" in relation_headers
    workbook.close()
    output.unlink()
    print("OK: linear and non-linear time-area relationships are classified and exported")


if __name__ == "__main__":
    main()
