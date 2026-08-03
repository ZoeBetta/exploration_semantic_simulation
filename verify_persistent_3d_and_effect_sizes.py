"""Regression checks for persistent 3-D preference and statistical effect sizes."""

from __future__ import annotations

import math
from pathlib import Path
from types import SimpleNamespace

from gui import RunViewer
from statistical_analysis import analyze_results, write_results_excel


class FakeCamera:
    """Minimal CameraHandle replacement that records scene open/close calls."""

    def __init__(self):
        self.active = False
        self.open_calls = []
        self.close_calls = 0

    def open(self, floors, pose):
        self.active = True
        self.open_calls.append((floors, pose))

    def close(self):
        self.active = False
        self.close_calls += 1


def _viewer(label: str, floors_marker: object) -> RunViewer:
    robot = SimpleNamespace(floor=0)
    state = SimpleNamespace(floors=floors_marker, robot=robot)
    viewer = RunViewer(state, label)
    viewer._camera3d = FakeCamera()
    return viewer


def check_camera_preference_survives_new_building() -> None:
    RunViewer._persistent_camera_enabled = False
    RunViewer._persistent_max_turbo = False
    RunViewer._restore_camera_after_turbo = False

    first = _viewer("building 1", ["floor-set-1"])
    first._toggle_camera3d((1.0, 2.0, 0.25))
    assert RunViewer._persistent_camera_enabled
    assert first._camera3d.active
    assert first._camera3d.open_calls[-1][0] == ["floor-set-1"]

    # Ending an episode closes only its process, not the user's preference.
    first._camera3d.close()
    assert RunViewer._persistent_camera_enabled

    second = _viewer("building 2", ["floor-set-2"])
    second._open_persistent_camera_if_requested((3.0, 4.0, 1.25))
    assert second._camera3d.active
    assert second._camera3d.open_calls[-1][0] == ["floor-set-2"]

    # Clicking the button again disables the preference for all later runs.
    second._toggle_camera3d((3.0, 4.0, 1.25))
    assert not RunViewer._persistent_camera_enabled
    third = _viewer("building 3", ["floor-set-3"])
    third._open_persistent_camera_if_requested((0.0, 0.0, 0.0))
    assert not third._camera3d.active


def _dummy_results() -> list[dict]:
    rows = []
    # Four observations per condition, with deliberately separated means and
    # non-zero within-group variance so all standardized effects are finite.
    for run, jitter in enumerate((-1.5, -0.5, 0.5, 1.5), start=1):
        for fw in (0.0, 1.0):
            for opt in (False, True):
                shift = 8.0 * fw + (5.0 if opt else 0.0)
                base = 30.0 + shift + jitter
                rows.append({
                    "run": run,
                    "start_attempt": 0,
                    "diagnostic_restarts": 0,
                    "Fw": fw,
                    "Opt": opt,
                    "Iw": 1.0,
                    "Dw": 1.0,
                    "Vw": 0.3,
                    "Cw": 3.0,
                    "Ow": 8.0,
                    "Wp": 24.0,
                    "Ms": 1.0,
                    "layout_mode": "office",
                    "include_objects": True,
                    "object_density": 1.0,
                    "Af": base,
                    "Af_star": base + 2.0,
                    "Atot": base - 3.0,
                    "Vf": 2,
                    "Cf": 1,
                    "texpl": 600.0,
                    "TP": 1,
                    "FP": 0,
                    "FN": 0,
                    "TN": 10.0,
                    "SAR_Sensitivity": 1.0,
                    "SAR_Specificity": 1.0,
                    "SAR_BalancedAccuracy": 1.0,
                    "SAR_MCC": 1.0,
                })
    return rows


def check_effect_sizes_and_excel() -> None:
    results = _dummy_results()
    analysis = analyze_results(results)

    for row in analysis["anova_all_conditions"] + analysis["anova_fw"]:
        assert row["effect_size_type"] == "eta_squared"
        assert math.isfinite(row["effect_size"])
        assert 0.0 <= row["effect_size"] <= 1.0
        assert row["effect_magnitude"] in {"trascurabile", "basso", "medio", "alto"}

    pairwise = analysis["tukey_all_conditions"] + analysis["tukey_fw"]
    assert pairwise
    for row in pairwise:
        assert row["effect_size_type"] == "hedges_g"
        assert math.isfinite(row["effect_size"])
        assert row["effect_magnitude"] in {"trascurabile", "basso", "medio", "alto"}

    for row in analysis["welch_opt"]:
        assert row["effect_size_type"] == "hedges_g"
        assert math.isfinite(row["effect_size"])
        assert row["effect_magnitude"] in {"trascurabile", "basso", "medio", "alto"}

    summary = []
    for fw in (0.0, 1.0):
        for opt in (False, True):
            group = [r for r in results if r["Fw"] == fw and r["Opt"] == opt]
            summary.append({
                "Fw": fw,
                "Opt": opt,
                "N": len(group),
                "Af_mean": sum(r["Af"] for r in group) / len(group),
                "Af_std": 0.0,
                "Af_star_mean": sum(r["Af_star"] for r in group) / len(group),
                "Af_star_std": 0.0,
                "Atot_mean": sum(r["Atot"] for r in group) / len(group),
                "Atot_std": 0.0,
            })

    output = Path("_verify_v35_effect_sizes.xlsx")
    write_results_excel(output, results, summary, analysis)
    assert output.exists() and output.stat().st_size > 0

    from openpyxl import load_workbook
    workbook = load_workbook(output, read_only=True)
    for sheet_name in ("ANOVA conditions", "Tukey conditions", "Welch Opt", "ANOVA Fw", "Tukey Fw"):
        headers = [cell.value for cell in next(workbook[sheet_name].iter_rows())]
        assert "effect_size" in headers
        assert "effect_magnitude" in headers
    workbook.close()
    output.unlink()


if __name__ == "__main__":
    check_camera_preference_survives_new_building()
    check_effect_sizes_and_excel()
    print("OK: 3-D preference persists and effect sizes are exported correctly.")
