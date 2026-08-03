"""Statistical analysis and Excel export for SAR exploration batches.

The simulator produces one row per episode.  This module adds three families
of group analyses for the coverage metrics Af, Af_star and Atot:

1. one-way ANOVA over all (Fw, Opt) conditions + Tukey-Kramer pairwise tests
   with Bonferroni correction;
2. Welch t-test pooling all Opt=ON episodes versus all Opt=OFF episodes;
3. one-way ANOVA pooling all episodes with the same Fw + Tukey-Kramer pairwise
   tests with Bonferroni correction.

Every inferential result also reports an effect size and a qualitative
interpretation: eta-squared for omnibus ANOVAs and Hedges' g for two-group or
pairwise comparisons.

It also studies the relationship between effective exploration time ``texpl``
(which excludes the explicit floor-change cost ``Ts``) and total explored area
``Atot``.  Pearson/Spearman association, linear regression and several
non-linear alternatives are compared using small-sample corrected AIC (AICc),
so the output states whether a linear description is adequate and, otherwise,
which curve family best describes the observed data.

The functions are intentionally independent from the live simulator so they can
also be reused from notebooks or command-line scripts reading the CSV output.
"""

from __future__ import annotations

# -----------------------------------------------------------------------------
# CODE-REVIEW NOTES
# Purpose: Condition summaries, inferential tests, multiple-comparison correction, console reports, and Excel worksheets.
# Coordinate convention: array indices are (row=y, column=x), while
# metric positions and GUI points are written as (x, y). Distances are
# metres unless a name explicitly ends in ``_cells`` or ``_px``.
# Reproducibility: stochastic operations must use the seeded RNG passed
# by the run; avoid module-global random calls in simulation logic.
# Separation of concerns: ground truth, robot-built maps, planner state,
# and rendering overlays are intentionally distinct representations.
# When modifying this file, preserve those boundaries and update the
# corresponding ``verify_*.py`` regression test.
# -----------------------------------------------------------------------------

import math
from collections import defaultdict
from pathlib import Path
from typing import Iterable

import numpy as np
from scipy import optimize, stats
from i18n import get_language, localize_analysis_text

METRICS = ("Af", "Af_star", "Atot")
ALPHA = 0.05
TIME_AREA_MIN_N = 5
TIME_AREA_NONLINEAR_DELTA_AICC = 2.0


def _finite_metric_values(rows: Iterable[dict], metric: str) -> np.ndarray:
    values = []
    for row in rows:
        try:
            value = float(row[metric])
        except (KeyError, TypeError, ValueError):
            continue
        if math.isfinite(value):
            values.append(value)
    return np.asarray(values, dtype=float)


def _condition_label(row: dict) -> str:
    fw = float(row["Fw"])
    opt = "ON" if bool(row["Opt"]) else "OFF"
    return f"Fw={fw:g}; Opt={opt}"


def _fw_label(row: dict) -> str:
    return f"Fw={float(row['Fw']):g}"


def _group_rows(rows: list[dict], key_fn) -> dict[str, list[dict]]:
    groups: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        groups[key_fn(row)].append(row)
    return dict(sorted(groups.items(), key=lambda item: item[0]))


def _sample_summary(values: np.ndarray) -> dict:
    n = int(values.size)
    if n == 0:
        return {"N": 0, "mean": math.nan, "std": math.nan}
    return {
        "N": n,
        "mean": float(np.mean(values)),
        "std": float(np.std(values, ddof=1)) if n > 1 else 0.0,
    }


def _effect_magnitude_eta_squared(value: float) -> str:
    """Interpret eta-squared using conventional behavioural-science cutoffs."""
    if not math.isfinite(value):
        return "non disponibile"
    value = abs(float(value))
    if value < 0.01:
        return "trascurabile"
    if value < 0.06:
        return "basso"
    if value < 0.14:
        return "medio"
    return "alto"


def _effect_magnitude_hedges_g(value: float) -> str:
    """Interpret the absolute standardized mean difference (Hedges' g)."""
    if not math.isfinite(value):
        return "non disponibile"
    value = abs(float(value))
    if value < 0.20:
        return "trascurabile"
    if value < 0.50:
        return "basso"
    if value < 0.80:
        return "medio"
    return "alto"


def _eta_squared(arrays: dict[str, np.ndarray]) -> float:
    """Return one-way ANOVA eta-squared = SS_between / SS_total."""
    nonempty = [values for values in arrays.values() if values.size > 0]
    if len(nonempty) < 2:
        return math.nan
    total_n = sum(values.size for values in nonempty)
    if total_n == 0:
        return math.nan
    grand_mean = float(sum(np.sum(values) for values in nonempty) / total_n)
    ss_between = float(sum(
        values.size * (float(np.mean(values)) - grand_mean) ** 2
        for values in nonempty
    ))
    ss_total = float(sum(np.sum((values - grand_mean) ** 2) for values in nonempty))
    if ss_total <= 0.0:
        return 0.0
    return max(0.0, min(1.0, ss_between / ss_total))


def _hedges_g(values_a: np.ndarray, values_b: np.ndarray) -> float:
    """Bias-corrected standardized mean difference for two independent groups.

    The sign follows ``mean(values_a) - mean(values_b)``.  The statistic is
    reported alongside Welch's t test as a descriptive standardized difference;
    its pooled standardizer does not change the unequal-variance hypothesis test.
    """
    n_a = int(values_a.size)
    n_b = int(values_b.size)
    if n_a < 2 or n_b < 2:
        return math.nan
    df = n_a + n_b - 2
    var_a = float(np.var(values_a, ddof=1))
    var_b = float(np.var(values_b, ddof=1))
    pooled_var = ((n_a - 1) * var_a + (n_b - 1) * var_b) / df
    diff = float(np.mean(values_a) - np.mean(values_b))
    if pooled_var <= 0.0:
        return 0.0 if diff == 0.0 else math.nan
    cohen_d = diff / math.sqrt(pooled_var)
    # Standard small-sample correction; extremely accurate for df >= 2 and
    # avoids adding another special-function dependency.
    correction = 1.0 - 3.0 / (4.0 * df - 1.0) if df > 1 else 1.0
    return float(correction * cohen_d)


def _anova(groups: dict[str, list[dict]], metric: str) -> dict:
    arrays = {
        label: _finite_metric_values(group, metric)
        for label, group in groups.items()
    }
    arrays = {label: values for label, values in arrays.items()
              if values.size > 0}
    k = len(arrays)
    total_n = int(sum(values.size for values in arrays.values()))
    eta_squared = _eta_squared(arrays)
    base = {
        "metric": metric,
        "groups": k,
        "N": total_n,
        "df_between": max(0, k - 1),
        "df_within": max(0, total_n - k),
        "effect_size_type": "eta_squared",
        "effect_size": eta_squared,
        "effect_magnitude": _effect_magnitude_eta_squared(eta_squared),
    }
    if k < 2 or total_n <= k:
        return {
            **base,
            "F": math.nan,
            "p": math.nan,
            "note": "dati insufficienti per ANOVA",
        }
    try:
        f_stat, p_value = stats.f_oneway(*arrays.values())
        note = ""
    except Exception as exc:  # pragma: no cover - defensive runtime path
        f_stat, p_value = math.nan, math.nan
        note = f"ANOVA non calcolabile: {exc}"
    return {
        **base,
        "df_between": k - 1,
        "df_within": total_n - k,
        "F": float(f_stat),
        "p": float(p_value),
        "note": note,
    }


def _tukey_kramer_bonferroni(groups: dict[str, list[dict]], metric: str,
                             alpha: float = ALPHA) -> list[dict]:
    arrays = {
        label: _finite_metric_values(group, metric)
        for label, group in groups.items()
    }
    arrays = {label: values for label, values in arrays.items()
              if values.size > 0}
    labels = sorted(arrays)
    k = len(labels)
    total_n = int(sum(arrays[label].size for label in labels))
    m = k * (k - 1) // 2
    if k < 2 or total_n <= k or m == 0:
        return []

    means = {label: float(np.mean(arrays[label])) for label in labels}
    ns = {label: int(arrays[label].size) for label in labels}
    sse = 0.0
    for label in labels:
        values = arrays[label]
        sse += float(np.sum((values - means[label]) ** 2))
    df_error = total_n - k
    mse = sse / df_error if df_error > 0 else math.nan

    rows = []
    for i, label_a in enumerate(labels):
        for label_b in labels[i + 1:]:
            diff = means[label_a] - means[label_b]
            se = math.sqrt(mse * 0.5 * (1.0 / ns[label_a] + 1.0 / ns[label_b])) if mse > 0 else 0.0
            if se > 0.0 and math.isfinite(se):
                q_stat = abs(diff) / se
                p_tukey = float(stats.studentized_range.sf(q_stat, k, df_error))
            elif diff == 0.0:
                q_stat = 0.0
                p_tukey = 1.0
            else:
                q_stat = math.inf
                p_tukey = 0.0
            p_bonf = min(1.0, p_tukey * m) if math.isfinite(p_tukey) else math.nan
            hedges_g = _hedges_g(arrays[label_a], arrays[label_b])
            rows.append({
                "metric": metric,
                "group_1": label_a,
                "group_2": label_b,
                "N_1": ns[label_a],
                "N_2": ns[label_b],
                "mean_1": means[label_a],
                "mean_2": means[label_b],
                "diff_1_minus_2": diff,
                "q": float(q_stat),
                "p_tukey": p_tukey,
                "p_bonferroni": p_bonf,
                "reject_0.05_bonferroni": bool(math.isfinite(p_bonf) and p_bonf < alpha),
                "effect_size_type": "hedges_g",
                "effect_size": hedges_g,
                "effect_magnitude": _effect_magnitude_hedges_g(hedges_g),
            })
    return rows


def _welch_opt(rows: list[dict], metric: str, alpha: float = ALPHA) -> dict:
    on_values = _finite_metric_values(
        [row for row in rows if bool(row.get("Opt"))], metric)
    off_values = _finite_metric_values(
        [row for row in rows if not bool(row.get("Opt"))], metric)
    on_summary = _sample_summary(on_values)
    off_summary = _sample_summary(off_values)
    hedges_g = _hedges_g(on_values, off_values)
    if on_values.size < 2 or off_values.size < 2:
        t_stat = p_value = math.nan
        note = "servono almeno due osservazioni per ciascun gruppo Opt"
    else:
        t_stat, p_value = stats.ttest_ind(
            on_values, off_values, equal_var=False, nan_policy="omit")
        note = "Welch t-test, varianze non assunte uguali"
    return {
        "metric": metric,
        "N_ON": on_summary["N"],
        "mean_ON": on_summary["mean"],
        "std_ON": on_summary["std"],
        "N_OFF": off_summary["N"],
        "mean_OFF": off_summary["mean"],
        "std_OFF": off_summary["std"],
        "diff_ON_minus_OFF": (
            on_summary["mean"] - off_summary["mean"]
            if math.isfinite(on_summary["mean"]) and math.isfinite(off_summary["mean"])
            else math.nan
        ),
        "t": float(t_stat),
        "p": float(p_value),
        "reject_0.05": bool(math.isfinite(float(p_value)) and float(p_value) < alpha),
        "effect_size_type": "hedges_g",
        "effect_size": hedges_g,
        "effect_magnitude": _effect_magnitude_hedges_g(hedges_g),
        "note": note,
    }


# ---------------------------------------------------------------------------
# Effective exploration time versus total explored area
# ---------------------------------------------------------------------------
def _finite_xy(rows: Iterable[dict], x_metric: str,
               y_metric: str) -> tuple[np.ndarray, np.ndarray]:
    """Return paired finite observations without silently breaking pairing."""
    x_values = []
    y_values = []
    for row in rows:
        try:
            x_value = float(row[x_metric])
            y_value = float(row[y_metric])
        except (KeyError, TypeError, ValueError):
            continue
        if math.isfinite(x_value) and math.isfinite(y_value):
            x_values.append(x_value)
            y_values.append(y_value)
    return (np.asarray(x_values, dtype=float),
            np.asarray(y_values, dtype=float))


def _aicc(rss: float, n: int, parameter_count: int) -> float:
    """Small-sample corrected Akaike information criterion for least squares."""
    if n <= parameter_count + 1 or not math.isfinite(rss):
        return math.inf
    # A perfect deterministic fit has RSS=0.  A tiny floor keeps the criterion
    # finite while preserving its overwhelming preference over imperfect fits.
    safe_rss = max(float(rss), np.finfo(float).tiny)
    aic = n * math.log(safe_rss / n) + 2.0 * parameter_count
    return float(
        aic
        + (2.0 * parameter_count * (parameter_count + 1))
        / (n - parameter_count - 1)
    )


def _ols_fit(x: np.ndarray, y: np.ndarray, model: str,
             design: np.ndarray, formula: str) -> dict:
    """Fit one candidate model and return comparable diagnostics."""
    n = int(y.size)
    k = int(design.shape[1])
    rank = int(np.linalg.matrix_rank(design))
    if n < k or rank < k:
        return {
            "model": model,
            "formula": formula,
            "N": n,
            "parameters": k,
            "rss": math.nan,
            "rmse": math.nan,
            "r2": math.nan,
            "adjusted_r2": math.nan,
            "aicc": math.inf,
            "intercept": math.nan,
            "coefficient_1": math.nan,
            "coefficient_2": math.nan,
            "coefficient_1_p": math.nan,
            "coefficient_2_p": math.nan,
            "tau": math.nan,
            "asymptote": math.nan,
            "note": "matrice del modello non identificabile",
        }

    coefficients, *_ = np.linalg.lstsq(design, y, rcond=None)
    fitted = design @ coefficients
    residuals = y - fitted
    rss = float(np.sum(residuals ** 2))
    tss = float(np.sum((y - float(np.mean(y))) ** 2))
    r2 = (1.0 - rss / tss) if tss > 0.0 else (1.0 if rss <= 1e-15 else 0.0)
    df_residual = n - k
    adjusted_r2 = (
        1.0 - (1.0 - r2) * (n - 1) / df_residual
        if df_residual > 0 else math.nan
    )
    rmse = math.sqrt(rss / n) if n > 0 else math.nan

    p_values = np.full(k, np.nan, dtype=float)
    if df_residual > 0:
        sigma_squared = rss / df_residual
        covariance = sigma_squared * np.linalg.pinv(design.T @ design)
        standard_errors = np.sqrt(np.maximum(np.diag(covariance), 0.0))
        with np.errstate(divide="ignore", invalid="ignore"):
            t_values = coefficients / standard_errors
        p_values = 2.0 * stats.t.sf(np.abs(t_values), df_residual)
        p_values = np.where(
            (standard_errors == 0.0) & (coefficients == 0.0), 1.0, p_values)

    return {
        "model": model,
        "formula": formula,
        "N": n,
        "parameters": k,
        "rss": rss,
        "rmse": float(rmse),
        "r2": float(r2),
        "adjusted_r2": float(adjusted_r2),
        "aicc": _aicc(rss, n, k),
        "intercept": float(coefficients[0]),
        "coefficient_1": (
            float(coefficients[1]) if coefficients.size > 1 else math.nan),
        "coefficient_2": (
            float(coefficients[2]) if coefficients.size > 2 else math.nan),
        "coefficient_1_p": (
            float(p_values[1]) if p_values.size > 1 else math.nan),
        "coefficient_2_p": (
            float(p_values[2]) if p_values.size > 2 else math.nan),
        "tau": math.nan,
        "asymptote": math.nan,
        "note": "",
    }


def _saturating_exponential_fit(x: np.ndarray, y: np.ndarray) -> dict:
    """Fit y = b0 + span * (1-exp(-x/tau)), with non-negative span."""
    n = int(y.size)
    k = 3
    base = {
        "model": "saturating_exponential",
        "formula": "Atot = b0 + span * (1 - exp(-texpl/tau))",
        "N": n,
        "parameters": k,
        "rss": math.nan,
        "rmse": math.nan,
        "r2": math.nan,
        "adjusted_r2": math.nan,
        "aicc": math.inf,
        "intercept": math.nan,
        "coefficient_1": math.nan,
        "coefficient_2": math.nan,
        "coefficient_1_p": math.nan,
        "coefficient_2_p": math.nan,
        "tau": math.nan,
        "asymptote": math.nan,
        "note": "",
    }
    if n <= k + 1 or np.unique(x).size < 4 or np.ptp(x) <= 0.0:
        return {**base, "note": "dati insufficienti per il modello saturante"}

    def model(values, intercept, span, tau):
        return intercept + span * (1.0 - np.exp(-values / tau))

    y_min = float(np.min(y))
    y_max = float(np.max(y))
    x_max = max(float(np.max(x)), 1.0)
    initial = [max(-20.0, min(100.0, y_min)),
               max(1e-3, y_max - y_min),
               max(1.0, float(np.median(x)))]
    lower = [-200.0, 0.0, 1e-6]
    upper = [200.0, 500.0, x_max * 1000.0 + 1.0]
    try:
        parameters, _ = optimize.curve_fit(
            model, x, y, p0=initial, bounds=(lower, upper), maxfev=30000)
        fitted = model(x, *parameters)
    except Exception as exc:
        return {**base, "note": f"modello saturante non calcolabile: {exc}"}

    residuals = y - fitted
    rss = float(np.sum(residuals ** 2))
    tss = float(np.sum((y - float(np.mean(y))) ** 2))
    r2 = (1.0 - rss / tss) if tss > 0.0 else (1.0 if rss <= 1e-15 else 0.0)
    df_residual = n - k
    adjusted_r2 = (
        1.0 - (1.0 - r2) * (n - 1) / df_residual
        if df_residual > 0 else math.nan
    )
    intercept, span, tau = (float(value) for value in parameters)
    return {
        **base,
        "rss": rss,
        "rmse": math.sqrt(rss / n),
        "r2": float(r2),
        "adjusted_r2": float(adjusted_r2),
        "aicc": _aicc(rss, n, k),
        "intercept": intercept,
        "coefficient_1": span,
        "tau": tau,
        "asymptote": intercept + span,
    }


def _relationship_strength(r2: float) -> str:
    if not math.isfinite(r2):
        return "non disponibile"
    if r2 < 0.25:
        return "debole"
    if r2 < 0.50:
        return "moderata"
    if r2 < 0.75:
        return "forte"
    return "molto forte"


def _time_area_model_label(model: dict) -> str:
    name = model.get("model")
    if name == "linear":
        return "lineare"
    if name == "quadratic":
        curvature = float(model.get("coefficient_2", math.nan))
        if math.isfinite(curvature) and curvature < 0.0:
            return "quadratica concava (rendimenti decrescenti)"
        if math.isfinite(curvature) and curvature > 0.0:
            return "quadratica convessa (crescita accelerata)"
        return "quadratica"
    if name in {"logarithmic", "square_root", "saturating_exponential"}:
        return "saturante / rendimenti decrescenti"
    return str(name or "non disponibile")


def _fit_time_area_scope(rows: list[dict], scope: str,
                         alpha: float = ALPHA) -> tuple[dict, list[dict]]:
    """Analyse one collection of episodes and compare plausible curve forms."""
    x, y = _finite_xy(rows, "texpl", "Atot")
    n = int(x.size)
    unique_x = int(np.unique(x).size)
    empty_summary = {
        "scope": scope,
        "N": n,
        "unique_times": unique_x,
        "pearson_r": math.nan,
        "pearson_p": math.nan,
        "spearman_rho": math.nan,
        "spearman_p": math.nan,
        "linear_slope": math.nan,
        "linear_slope_p": math.nan,
        "linear_r2": math.nan,
        "linear_adjusted_r2": math.nan,
        "best_model": "non_disponibile",
        "best_model_label": "non disponibile",
        "best_model_r2": math.nan,
        "best_model_adjusted_r2": math.nan,
        "delta_aicc_vs_linear": math.nan,
        "relationship": "non valutabile",
        "relationship_strength": "non disponibile",
        "nonlinearity_evidence": "non disponibile",
        "quadratic_term_p": math.nan,
        "note": "",
    }
    if n < TIME_AREA_MIN_N or unique_x < 3:
        reason = (
            f"servono almeno {TIME_AREA_MIN_N} osservazioni e 3 tempi distinti"
        )
        return {**empty_summary, "note": reason}, []

    if np.std(x) <= 0.0 or np.std(y) <= 0.0:
        return {
            **empty_summary,
            "note": "tempo o coverage senza variabilita; relazione non stimabile",
        }, []

    pearson_r, pearson_p = stats.pearsonr(x, y)
    spearman_rho, spearman_p = stats.spearmanr(x, y)

    models = [
        _ols_fit(
            x, y, "linear", np.column_stack((np.ones(n), x)),
            "Atot = b0 + b1 * texpl",
        ),
        _ols_fit(
            x, y, "quadratic",
            np.column_stack((np.ones(n), x, x ** 2)),
            "Atot = b0 + b1 * texpl + b2 * texpl^2",
        ),
        _ols_fit(
            x, y, "logarithmic",
            np.column_stack((np.ones(n), np.log1p(np.maximum(x, 0.0)))),
            "Atot = b0 + b1 * log(1 + texpl)",
        ),
        _ols_fit(
            x, y, "square_root",
            np.column_stack((np.ones(n), np.sqrt(np.maximum(x, 0.0)))),
            "Atot = b0 + b1 * sqrt(texpl)",
        ),
        _saturating_exponential_fit(x, y),
    ]
    valid_models = [model for model in models
                    if math.isfinite(float(model.get("aicc", math.inf)))]
    if not valid_models:
        return {**empty_summary, "note": "nessun modello stimabile"}, models

    valid_models.sort(key=lambda model: (model["aicc"], model["parameters"]))
    best = valid_models[0]
    min_aicc = float(best["aicc"])
    weights_raw = []
    for model in models:
        aicc = float(model.get("aicc", math.inf))
        delta = aicc - min_aicc if math.isfinite(aicc) else math.inf
        model["scope"] = scope
        model["delta_aicc"] = float(delta)
        weight_raw = math.exp(-0.5 * delta) if math.isfinite(delta) else 0.0
        weights_raw.append(weight_raw)
    weight_sum = sum(weights_raw)
    for model, weight_raw in zip(models, weights_raw):
        model["akaike_weight"] = (
            float(weight_raw / weight_sum) if weight_sum > 0.0 else math.nan)

    linear = next(model for model in models if model["model"] == "linear")
    quadratic = next(model for model in models if model["model"] == "quadratic")
    linear_aicc = float(linear["aicc"])
    improvement = (linear_aicc - min_aicc
                   if math.isfinite(linear_aicc) else math.nan)

    # Differences below 2 AICc units are not considered enough to reject a
    # simpler linear description.  This deliberately avoids labelling tiny,
    # noisy improvements as substantive non-linearity.
    approximately_linear = (
        best["model"] == "linear"
        or (math.isfinite(improvement)
            and improvement < TIME_AREA_NONLINEAR_DELTA_AICC)
    )
    if approximately_linear:
        relationship = "approssimativamente lineare"
        selected = linear
    else:
        relationship = f"non lineare: {_time_area_model_label(best)}"
        selected = best

    if not math.isfinite(improvement) or improvement < 2.0:
        evidence = "nessuna/trascurabile"
    elif improvement < 6.0:
        evidence = "debole"
    elif improvement < 10.0:
        evidence = "moderata"
    else:
        evidence = "forte"

    summary = {
        **empty_summary,
        "pearson_r": float(pearson_r),
        "pearson_p": float(pearson_p),
        "spearman_rho": float(spearman_rho),
        "spearman_p": float(spearman_p),
        "linear_slope": float(linear["coefficient_1"]),
        "linear_slope_p": float(linear["coefficient_1_p"]),
        "linear_r2": float(linear["r2"]),
        "linear_adjusted_r2": float(linear["adjusted_r2"]),
        "best_model": selected["model"],
        "best_model_label": _time_area_model_label(selected),
        "best_model_r2": float(selected["r2"]),
        "best_model_adjusted_r2": float(selected["adjusted_r2"]),
        "delta_aicc_vs_linear": float(max(0.0, improvement)),
        "relationship": relationship,
        "relationship_strength": _relationship_strength(float(selected["r2"])),
        "nonlinearity_evidence": evidence,
        "quadratic_term_p": float(quadratic["coefficient_2_p"]),
        "note": (
            "AICc confronta lineare, quadratica, logaritmica, radice quadrata "
            "ed esponenziale saturante; Delta AICc < 2 mantiene la descrizione "
            "lineare per parsimonia."
        ),
    }
    return summary, models


def _time_area_scopes(results: list[dict]) -> list[tuple[str, list[dict]]]:
    """Overall and stratified scopes, ordered for readable output."""
    scopes: list[tuple[str, list[dict]]] = [("Complessiva", results)]
    opt_groups = _group_rows(
        results,
        lambda row: f"Opt={'ON' if bool(row.get('Opt')) else 'OFF'}",
    )
    scopes.extend(opt_groups.items())
    fw_groups = _group_rows(results, _fw_label)
    scopes.extend(fw_groups.items())
    condition_groups = _group_rows(results, _condition_label)
    scopes.extend(condition_groups.items())
    return scopes


def _analyze_time_area_relationship(results: list[dict],
                                    alpha: float = ALPHA) -> tuple[list[dict], list[dict]]:
    summaries = []
    model_rows = []
    seen = set()
    for scope, rows in _time_area_scopes(results):
        # A single Fw or Opt value can make labels duplicate the complete set.
        # Retain each semantic scope once, but never duplicate an identical
        # label accidentally.
        if scope in seen:
            continue
        seen.add(scope)
        summary, models = _fit_time_area_scope(rows, scope, alpha)
        summaries.append(summary)
        model_rows.extend(models)
    return summaries, model_rows


def analyze_results(results: list[dict], alpha: float = ALPHA) -> dict:
    """Return all requested statistical analyses in serializable structures."""
    condition_groups = _group_rows(results, _condition_label)
    fw_groups = _group_rows(results, _fw_label)
    time_area_relationship, time_area_models = _analyze_time_area_relationship(
        results, alpha)

    return {
        "alpha": alpha,
        "metrics": list(METRICS),
        "anova_all_conditions": [
            _anova(condition_groups, metric) for metric in METRICS
        ],
        "tukey_all_conditions": [
            row for metric in METRICS
            for row in _tukey_kramer_bonferroni(condition_groups, metric, alpha)
        ],
        "welch_opt": [
            _welch_opt(results, metric, alpha) for metric in METRICS
        ],
        "anova_fw": [
            _anova(fw_groups, metric) for metric in METRICS
        ],
        "tukey_fw": [
            row for metric in METRICS
            for row in _tukey_kramer_bonferroni(fw_groups, metric, alpha)
        ],
        "time_area_relationship": time_area_relationship,
        "time_area_models": time_area_models,
    }


def _fmt(value, digits: int = 4) -> str:
    if isinstance(value, bool):
        return "SI" if value else "NO"
    try:
        f_value = float(value)
    except (TypeError, ValueError):
        return str(value)
    if not math.isfinite(f_value):
        return "-"
    return f"{f_value:.{digits}g}"


def print_statistical_analysis(analysis: dict) -> None:
    """Print all statistical results in the selected interface language."""
    english = get_language() == "en"
    print("\n" + "=" * 78)
    print("STATISTICAL TESTS FOR COVERAGE METRICS" if english
          else "TEST STATISTICI SUI PARAMETRI DI COVERAGE")
    print("Metrics: A_f, A_f*, A_tot. Alpha = 0.05" if english
          else "Metriche: A_f, A_f*, A_tot. Alpha = 0.05")

    print("\n1) ANOVA across all conditions (Fw, Opt)" if english
          else "\n1) ANOVA su tutte le condizioni (Fw, Opt)")
    for row in analysis["anova_all_conditions"]:
        note = row.get("note", "")
        print(
            f"  {row['metric']}: F({row['df_between']}, {row['df_within']})="
            f"{_fmt(row['F'])}, p={_fmt(row['p'])}, "
            f"eta^2={_fmt(row['effect_size'])} "
            f"({localize_analysis_text(row['effect_magnitude'])})"
            + (f" [{note}]" if note else "")
        )
    print("\n   Tukey-Kramer post-hoc tests with Bonferroni correction" if english
          else "\n   Tukey-Kramer post-hoc con correzione Bonferroni")
    for row in analysis["tukey_all_conditions"]:
        significant_label = "significant" if english else "significativo"
        print(
            f"  {row['metric']}: {row['group_1']} vs {row['group_2']} | "
            f"diff={_fmt(row['diff_1_minus_2'])}, "
            f"q={_fmt(row['q'])}, p_Tukey={_fmt(row['p_tukey'])}, "
            f"p_Bonf={_fmt(row['p_bonferroni'])}, "
            f"{significant_label}={_fmt(row['reject_0.05_bonferroni'])}, "
            f"Hedges g={_fmt(row['effect_size'])} "
            f"({localize_analysis_text(row['effect_magnitude'])})"
        )

    print("\n2) Analysis stratified by Opt: ON vs OFF, Welch t-test" if english
          else "\n2) Analisi stratificata per Opt: ON vs OFF, Welch t-test")
    for row in analysis["welch_opt"]:
        mean_label = "mean" if english else "media"
        significant_label = "significant" if english else "significativo"
        print(
            f"  {row['metric']}: ON n={row['N_ON']} {mean_label}={_fmt(row['mean_ON'])} "
            f"vs OFF n={row['N_OFF']} {mean_label}={_fmt(row['mean_OFF'])}; "
            f"diff={_fmt(row['diff_ON_minus_OFF'])}, "
            f"t={_fmt(row['t'])}, p={_fmt(row['p'])}, "
            f"{significant_label}={_fmt(row['reject_0.05'])}, "
            f"Hedges g={_fmt(row['effect_size'])} "
            f"({localize_analysis_text(row['effect_magnitude'])})"
        )

    print("\n3) Analysis stratified by Fw, pooling Opt: ANOVA" if english
          else "\n3) Analisi stratificata per Fw, aggregando Opt: ANOVA")
    for row in analysis["anova_fw"]:
        note = row.get("note", "")
        print(
            f"  {row['metric']}: F({row['df_between']}, {row['df_within']})="
            f"{_fmt(row['F'])}, p={_fmt(row['p'])}, "
            f"eta^2={_fmt(row['effect_size'])} "
            f"({localize_analysis_text(row['effect_magnitude'])})"
            + (f" [{note}]" if note else "")
        )
    print("\n   Tukey-Kramer post-hoc tests among Fw groups with Bonferroni correction"
          if english else
          "\n   Tukey-Kramer post-hoc fra gruppi Fw con correzione Bonferroni")
    for row in analysis["tukey_fw"]:
        significant_label = "significant" if english else "significativo"
        print(
            f"  {row['metric']}: {row['group_1']} vs {row['group_2']} | "
            f"diff={_fmt(row['diff_1_minus_2'])}, "
            f"q={_fmt(row['q'])}, p_Tukey={_fmt(row['p_tukey'])}, "
            f"p_Bonf={_fmt(row['p_bonferroni'])}, "
            f"{significant_label}={_fmt(row['reject_0.05_bonferroni'])}, "
            f"Hedges g={_fmt(row['effect_size'])} "
            f"({localize_analysis_text(row['effect_magnitude'])})"
        )

    print("\n4) Relationship between effective exploration time and A_tot"
          if english else
          "\n4) Relazione tra tempo effettivo di esplorazione e A_tot")
    print(
        "   t_expl is the simulated time effectively available for exploration: "
        "the Ts cost of floor transitions is removed from the budget but is not "
        "counted in t_expl."
        if english else
        "   t_expl è il tempo simulato realmente disponibile per esplorare: "
        "il costo Ts dei cambi di piano viene sottratto dal budget ma non "
        "conteggiato in t_expl."
    )
    for row in analysis.get("time_area_relationship", []):
        prefix = ("  OVERALL" if english else "  COMPLESSIVA")             if row["scope"] == "Complessiva" else f"  {row['scope']}"
        if row["relationship"] == "non valutabile":
            label = "not evaluable" if english else "non valutabile"
            print(f"{prefix}: {label} [{row.get('note', '')}]")
            continue
        r2_label = "linear R^2" if english else "lineare R^2"
        slope_label = "slope" if english else "pendenza"
        conclusion_label = "conclusion" if english else "conclusione"
        strength_label = "strength" if english else "forza"
        evidence_label = "non-linearity evidence" if english else "evidenza non-linearità"
        print(
            f"{prefix}: n={row['N']}, Pearson r={_fmt(row['pearson_r'])} "
            f"(p={_fmt(row['pearson_p'])}), "
            f"{r2_label}={_fmt(row['linear_r2'])}, "
            f"{slope_label}={_fmt(row['linear_slope'])} percentage points/s "
            f"(p={_fmt(row['linear_slope_p'])}); "
            f"{conclusion_label}={localize_analysis_text(row['relationship'])}, "
            f"{strength_label}={localize_analysis_text(row['relationship_strength'])}, "
            f"Delta AICc={_fmt(row['delta_aicc_vs_linear'])}, "
            f"{evidence_label}={localize_analysis_text(row['nonlinearity_evidence'])}"
        )
    print("=" * 78)


def _write_rows(sheet, rows: list[dict], headers: list[str]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def _style_sheet(sheet) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="4A6FA5")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D5DCE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.freeze_panes = "A2"

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = 0
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        width = min(max(10, max_len + 2), 42)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_results_excel(path: str | Path, results: list[dict],
                        condition_summary: list[dict], analysis: dict) -> None:
    """Write raw results, aggregate summaries and statistical tests to .xlsx."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency path
        raise RuntimeError(
            "Per esportare Excel installare openpyxl: python -m pip install openpyxl"
        ) from exc

    wb = Workbook()
    default = wb.active
    default.title = "Run results"

    run_headers = [
        "run", "start_attempt", "diagnostic_restarts", "Fw", "Opt",
        "Iw", "Dw", "Vw", "Cw", "Ow", "Wp", "Ms",
        "layout_mode", "include_objects", "object_density",
        "Af", "Af_star", "Atot", "Vf", "Cf", "texpl",
        "TP", "FP", "FN", "TN",
        "SAR_Sensitivity", "SAR_Specificity", "SAR_BalancedAccuracy", "SAR_MCC",
    ]
    _write_rows(default, results, run_headers)

    sheets = [default]

    condition_headers = [
        "Fw", "Opt", "N", "Af_mean", "Af_std",
        "Af_star_mean", "Af_star_std", "Atot_mean", "Atot_std",
    ]
    ws = wb.create_sheet("Condition stats")
    _write_rows(ws, condition_summary, condition_headers)
    sheets.append(ws)

    ws = wb.create_sheet("ANOVA conditions")
    _write_rows(ws, analysis["anova_all_conditions"], [
        "metric", "groups", "N", "df_between", "df_within", "F", "p",
        "effect_size_type", "effect_size", "effect_magnitude", "note"
    ])
    sheets.append(ws)

    ws = wb.create_sheet("Tukey conditions")
    _write_rows(ws, analysis["tukey_all_conditions"], [
        "metric", "group_1", "group_2", "N_1", "N_2", "mean_1", "mean_2",
        "diff_1_minus_2", "q", "p_tukey", "p_bonferroni", "reject_0.05_bonferroni",
        "effect_size_type", "effect_size", "effect_magnitude",
    ])
    sheets.append(ws)

    ws = wb.create_sheet("Welch Opt")
    _write_rows(ws, analysis["welch_opt"], [
        "metric", "N_ON", "mean_ON", "std_ON", "N_OFF", "mean_OFF", "std_OFF",
        "diff_ON_minus_OFF", "t", "p", "reject_0.05",
        "effect_size_type", "effect_size", "effect_magnitude", "note",
    ])
    sheets.append(ws)

    ws = wb.create_sheet("ANOVA Fw")
    _write_rows(ws, analysis["anova_fw"], [
        "metric", "groups", "N", "df_between", "df_within", "F", "p",
        "effect_size_type", "effect_size", "effect_magnitude", "note"
    ])
    sheets.append(ws)

    ws = wb.create_sheet("Tukey Fw")
    _write_rows(ws, analysis["tukey_fw"], [
        "metric", "group_1", "group_2", "N_1", "N_2", "mean_1", "mean_2",
        "diff_1_minus_2", "q", "p_tukey", "p_bonferroni", "reject_0.05_bonferroni",
        "effect_size_type", "effect_size", "effect_magnitude",
    ])
    sheets.append(ws)

    ws = wb.create_sheet("Time-area relation")
    _write_rows(ws, analysis.get("time_area_relationship", []), [
        "scope", "N", "unique_times",
        "pearson_r", "pearson_p", "spearman_rho", "spearman_p",
        "linear_slope", "linear_slope_p", "linear_r2",
        "linear_adjusted_r2", "best_model", "best_model_label",
        "best_model_r2", "best_model_adjusted_r2",
        "delta_aicc_vs_linear", "relationship", "relationship_strength",
        "nonlinearity_evidence", "quadratic_term_p", "note",
    ])
    sheets.append(ws)

    ws = wb.create_sheet("Time-area models")
    _write_rows(ws, analysis.get("time_area_models", []), [
        "scope", "model", "formula", "N", "parameters", "rss", "rmse",
        "r2", "adjusted_r2", "aicc", "delta_aicc", "akaike_weight",
        "intercept", "coefficient_1", "coefficient_2",
        "coefficient_1_p", "coefficient_2_p", "tau", "asymptote", "note",
    ])
    sheets.append(ws)

    notes = wb.create_sheet("Notes")
    notes.append(["Voce", "Descrizione"])
    notes.append(["ANOVA conditions", "One-way ANOVA sui gruppi definiti da ogni coppia (Fw, Opt)."])
    notes.append(["Tukey conditions", "Tukey-Kramer pairwise; p_bonferroni = p_Tukey moltiplicato per il numero di confronti."])
    notes.append(["Welch Opt", "T-test di Welch tra episodi Opt=ON e Opt=OFF, aggregando tutti i Fw."])
    notes.append(["ANOVA Fw", "One-way ANOVA tra gruppi Fw, aggregando Opt=ON e Opt=OFF."])
    notes.append(["Effect size ANOVA", "Eta-squared (eta^2): trascurabile <0.01; basso <0.06; medio <0.14; alto >=0.14."])
    notes.append(["Effect size confronti", "Hedges' g con segno secondo gruppo_1-gruppo_2 (o ON-OFF): trascurabile |g|<0.20; basso <0.50; medio <0.80; alto >=0.80."])
    notes.append(["Metriche", "Af, Af_star e Atot sono percentuali di coverage."])
    notes.append([
        "Tempo effettivo",
        "texpl e il tempo simulato realmente speso nell'esplorazione. Il costo Ts di ogni cambio di piano riduce il budget residuo, ma non incrementa texpl.",
    ])
    notes.append([
        "Time-area relation",
        "Relazione fra texpl e Atot. Riporta Pearson/Spearman, regressione lineare e selezione del modello con AICc.",
    ])
    notes.append([
        "Decisione linearita",
        "Sono confrontati modelli lineare, quadratico, logaritmico, radice quadrata ed esponenziale saturante. Se il migliore modello non lineare migliora AICc di meno di 2 punti, la relazione resta classificata come approssimativamente lineare per parsimonia.",
    ])
    notes.append([
        "Interpretazione curva",
        "Logaritmica, radice quadrata ed esponenziale saturante indicano rendimenti decrescenti; la quadratica e classificata come concava o convessa in base al segno del termine di secondo grado.",
    ])
    sheets.append(notes)

    for sheet in sheets:
        _style_sheet(sheet)

    path = Path(path)
    wb.save(path)
